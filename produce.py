#! /usr/bin/env python3

import openpyxl
import logging

logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')
wb = openpyxl.load_workbook("produceSales.xlsx")
sheets = wb.get_sheet_names()
font_s1 = openpyxl.styles.Font(
    b=True, color=openpyxl.styles.colors.GREEN, size=24)

logging.debug("Sheet Names: %s", sheets)
logging.debug("Open first sheet")
sheet1 = wb.get_sheet_by_name(sheets[0])
for i in range(2, sheet1.max_row + 1):
    if sheet1.cell(row=i, column=1).value == 'Celery':
        logging.debug("Found Celery @ row = %i", i)
        sheet1.cell(row=i, column=2).value = 1.19
    if sheet1.cell(row=i, column=1).value == 'Garlic':
        logging.debug("Found Garlic @ row = %i", i)
        sheet1.cell(row=i, column=2).value = 3.07
    if sheet1.cell(row=i, column=1).value == 'Lemon':
        logging.debug("Found Lemon @ row = %i", i)
        sheet1.cell(row=i, column=2).value = 1.27
wb.save("up.xlsx")
