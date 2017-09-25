#! /usr/bin/env python3

import openpyxl
import logging
import pprint


logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.get_active_sheet()
logging.debug('Max Rows: %s', sheet.max_row)
logging.debug('Max Cols: %s', sheet.max_column)
data = {}
for i in range(2, sheet.max_row + 1):
	state = sheet.cell(row=i, column=2).value
	county = sheet.cell(row=i, column=3).value
	population = sheet.cell(row=i, column=4).value
	data.setdefault(state, {})
	data[state].setdefault(county, {'tracts': 0, 'pop': 0})
	data[state][county]['tracts'] += 1
	data[state][county]['pop'] += int(population)

pprint.pprint(data)
