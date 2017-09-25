#! /usr/bin/env python3

import openpyxl
import logging


logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_active_sheet()
logging.debug('Max Cols: ' + str(sheet.max_column))
logging.debug('Max Rows: ' + str(sheet.max_row))
logging.debug(openpyxl.utils.get_column_letter(27))
logging.debug(openpyxl.utils.column_index_from_string('AB'))

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print('Value of ' + str(sheet.cell(row=i, column=j).coordinate) +
              ' is: ' + str(sheet.cell(row=i, column=j).value))
#        print('Value of ' + openpyxl.utils.get_column_letter(j) +
#              str(i) + ' is: ' + str(sheet.cell(row=i, column=j).value))
